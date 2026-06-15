"""
Financial report parsers — one per platform.

Usage:
    from app.financials.parsers import parse_tiktok, parse_lazada, parse_shopee

Each returns List[dict] where every dict maps to FinancialTransaction fields.
"""
import re
from datetime import date, datetime
from typing import List, Dict, Any

import openpyxl


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _to_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%Y/%m/%d", "%d %b %Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _month_str(d: date | None) -> str:
    if d is None:
        return "unknown"
    return d.strftime("%Y-%m")


# ─────────────────────────────────────────────────────────────────────────────
# TikTok Shop parser
# ─────────────────────────────────────────────────────────────────────────────

def _tiktok_scraper_singleton():
    """Return a cached TikTokScraper instance (loads session from disk)."""
    if not hasattr(_tiktok_scraper_singleton, "_instance"):
        try:
            from app.scrapers.tiktok import TikTokScraper
            _tiktok_scraper_singleton._instance = TikTokScraper()
        except Exception:
            _tiktok_scraper_singleton._instance = None
    return _tiktok_scraper_singleton._instance


def _shopee_client_singleton(shop_key: str = "my"):
    """Return a cached ShopeeAPIClient for a configured Shopee shop key (my/sg)."""
    if not hasattr(_shopee_client_singleton, "_instances"):
        _shopee_client_singleton._instances = {}
    instances = _shopee_client_singleton._instances
    if shop_key in instances:
        return instances[shop_key]

    try:
        from app.scrapers.shopee_api import ShopeeAPIClient, SHOPS
        shop_id = (SHOPS.get(shop_key) or {}).get("shop_id")
        if not shop_id:
            instances[shop_key] = None
        else:
            instances[shop_key] = ShopeeAPIClient(shop_id)
    except Exception as e:
        print(f"[Financials] Could not create ShopeeAPIClient ({shop_key}): {e}")
        instances[shop_key] = None
    return instances[shop_key]


def _normalize_order_id(value: Any) -> str:
    """
    Normalize order IDs read from Excel/API text.
    Converts numeric-like values such as 241234567890123.0 -> 241234567890123.
    """
    s = str(value or "").strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return s


def _shopee_fetch_order_qtys(order_ids: List[str]) -> Dict[str, int]:
    """
    Batch-fetch order quantities from Shopee API.
    Returns dict of {order_id: total_qty}.
    """
    order_ids = [_normalize_order_id(oid) for oid in order_ids]
    order_ids = [oid for oid in order_ids if oid]
    if not order_ids:
        return {}

    my_client = _shopee_client_singleton("my")
    sg_client = _shopee_client_singleton("sg")
    if not my_client and not sg_client:
        return {}

    qtys: Dict[str, int] = {}

    def _fetch_with_client(client, unresolved_ids: List[str]):
        if not client or not unresolved_ids:
            return
        # API allows max 50 per call
        for i in range(0, len(unresolved_ids), 50):
            batch = unresolved_ids[i:i + 50]
            try:
                resp = client.get_order_detail(batch)
                for order in resp.get("response", {}).get("order_list", []):
                    oid = _normalize_order_id(order.get("order_sn", ""))
                    items = order.get("item_list", [])
                    total_qty = sum(int(item.get("model_quantity_purchased", 1)) for item in items)
                    if oid:
                        qtys[oid] = max(total_qty, 1)
            except Exception as e:
                print(f"[Financials] Shopee order detail API error: {e}")

    # First try MY shop, then SG for unresolved order IDs.
    _fetch_with_client(my_client, order_ids)
    unresolved = [oid for oid in order_ids if oid not in qtys]
    if unresolved:
        _fetch_with_client(sg_client, unresolved)

    return qtys


def parse_tiktok(filepath: str, upload_batch: str = "") -> List[Dict[str, Any]]:
    """
    Sheet: 'Order details'
    Header row: 0  |  Data starts: row 1
    Key columns:
        Order/adjustment ID, Type, Order created time, Order settled time,
        Total settlement amount, Total Revenue, Total Fees,
        Transaction fee, TikTok Shop commission fee,
        Seller shipping fee, Actual shipping fee,
        Platform shipping fee discount, Affiliate Commission,
        Platform support fee, Seller co-funded voucher discount,
        Customer payment
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Order details"]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    def g(row, name):
        idx = col.get(name.strip())
        return row[idx] if idx is not None and idx < len(row) else None

    results = []
    for row in rows[1:]:
        if not any(row):
            continue
        order_type = str(g(row, "Type ") or "").strip()
        if order_type.lower() not in ("order", "adjustment", ""):
            continue

        _oid_raw = g(row, "Order/adjustment ID  ")
        if _oid_raw is None:
            continue
        # Excel may read numeric IDs as floats (e.g. 702228997654.0) — strip decimal
        try:
            order_id = str(int(float(str(_oid_raw).strip())))
        except (ValueError, OverflowError):
            order_id = str(_oid_raw).strip()
        if not order_id:
            continue
        order_date    = _to_date(g(row, "Order created time"))
        settle_date   = _to_date(g(row, "Order settled time"))
        d             = settle_date or order_date

        gross_rev     = _to_float(g(row, "Total Revenue"))
        customer_paid = _to_float(g(row, "Customer payment"))
        total_fees    = _to_float(g(row, "Total Fees"))
        commission    = _to_float(g(row, "TikTok Shop commission fee"))
        txn_fee       = _to_float(g(row, "Transaction fee"))
        platform_fee  = _to_float(g(row, "Platform support fee"))
        net_settle    = _to_float(g(row, "Total settlement amount"))

        ship_seller   = _to_float(g(row, "Seller shipping fee"))
        ship_actual   = _to_float(g(row, "Actual shipping fee"))
        ship_discount = _to_float(g(row, "Platform shipping fee discount"))
        ship_customer = _to_float(g(row, "Customer shipping fee"))
        net_ship      = ship_seller + ship_actual + ship_discount + ship_customer

        voucher_seller = _to_float(g(row, "Seller co-funded voucher discount"))
        affiliate      = _to_float(g(row, "Affiliate Commission"))

        other_fees = total_fees - commission - txn_fee - platform_fee

        # Parse qty and first product_id from "product_id * N; product_id * M;" format
        items_raw = str(g(row, "Shopping center items") or "").strip()
        qty = 0
        first_product_id = None
        api_product_name = None
        if items_raw and items_raw != "/":
            for m in re.finditer(r'(\d+)\s*\*\s*(\d+)', items_raw):
                if first_product_id is None:
                    first_product_id = m.group(1)
                qty += int(m.group(2))
        else:
            # Shopping center items is '/' or empty — call TikTok API for order details
            scraper = _tiktok_scraper_singleton()
            if scraper:
                info = scraper.search_order_sku_sync(order_id)
                if info.get("sku"):
                    first_product_id = info["sku"]
                    api_product_name = info.get("product_name") or None
                    qty = info.get("qty") or 1
        qty = max(qty, 1)

        results.append({
            "platform":        "tiktok",
            "product_name":    api_product_name,
            "sku":             first_product_id,  # TikTok SKU id — further resolved in _enrich_tiktok
            "qty":             qty,
            "order_id":        order_id,
            "order_date":      order_date,
            "settlement_date": settle_date,
            "month":           _month_str(d),
            "year":            d.year if d else 0,
            "gross_revenue":   gross_rev,
            "customer_paid":   customer_paid,
            "commission_fee":  commission,
            "transaction_fee": txn_fee,
            "service_fee":     platform_fee,
            "other_fees":      affiliate + other_fees,
            "total_fees":      total_fees,
            "shipping_buyer":  ship_customer,
            "shipping_cost":   ship_actual,
            "shipping_rebate": ship_discount,
            "net_shipping":    net_ship,
            "voucher_seller":  voucher_seller,
            "voucher_platform":0.0,
            "net_settlement":  net_settle,
            "upload_batch":    upload_batch,
        })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Lazada parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_lazada(filepath: str, upload_batch: str = "") -> List[Dict[str, Any]]:
    """
    Sheet: 'Income Overview'
    Header row: 0  |  Data starts: row 1
    One ROW per fee line — must group by Order Number.

    Fee Name examples:
        'Item Price Credit'          → revenue (positive)
        'Payment Fee'                → commission/txn (negative)
        'LazCoins Discount Promotion Fee' → platform voucher (negative, but platform-funded)
        'Shipping Fee'               → shipping deduction (negative)
        'Shipping Subsidy Credit'    → shipping rebate (positive)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Income Overview"]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    def g(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    # Group by order number
    orders: Dict[str, dict] = {}

    for row in rows[1:]:
        if not any(row):
            continue
        order_id = str(g(row, "Order Number") or "").strip()
        if not order_id:
            continue

        fee_name   = str(g(row, "Fee Name") or "").strip()
        amount     = _to_float(g(row, "Amount(Include Tax)"))
        txn_date   = _to_date(g(row, "Transaction Date"))
        release_dt = _to_date(g(row, "Release Date"))
        status     = str(g(row, "Release Status") or "").strip()

        product_name = str(g(row, "Product Name") or "").strip()
        seller_sku   = str(g(row, "Seller SKU") or "").strip()

        if order_id not in orders:
            orders[order_id] = {
                "platform":        "lazada",
                "order_id":        order_id,
                "order_date":      _to_date(g(row, "Order Creation Date")),
                "settlement_date": release_dt,
                "product_name":    product_name or None,
                "sku":             seller_sku or None,
                "qty":             1,
                "gross_revenue":   0.0,
                "customer_paid":   0.0,
                "commission_fee":  0.0,
                "transaction_fee": 0.0,
                "service_fee":     0.0,
                "other_fees":      0.0,
                "total_fees":      0.0,
                "shipping_buyer":  0.0,
                "shipping_cost":   0.0,
                "shipping_rebate": 0.0,
                "net_shipping":    0.0,
                "voucher_seller":  0.0,
                "voucher_platform":0.0,
                "net_settlement":  0.0,
                "upload_batch":    upload_batch,
            }

        o = orders[order_id]
        fn_lower = fee_name.lower()

        if "item price credit" in fn_lower:
            o["gross_revenue"] += amount
        elif "payment fee" in fn_lower:
            o["commission_fee"] += amount          # negative
        elif "shipping fee credit" in fn_lower or "shipping subsidy" in fn_lower:
            o["shipping_rebate"] += amount
        elif "shipping fee" in fn_lower and amount < 0:
            o["shipping_cost"] += amount
        elif "lazcoins" in fn_lower or "voucher" in fn_lower or "discount promotion" in fn_lower:
            o["voucher_platform"] += amount        # platform-funded
        elif "seller voucher" in fn_lower or "seller discount" in fn_lower:
            o["voucher_seller"] += amount
        else:
            o["other_fees"] += amount

        # update settlement date to latest release date seen
        if release_dt and (o["settlement_date"] is None or release_dt > o["settlement_date"]):
            o["settlement_date"] = release_dt

    # Post-process each order
    results = []
    for o in orders.values():
        o["total_fees"]   = o["commission_fee"] + o["transaction_fee"] + o["service_fee"] + o["other_fees"]
        o["net_shipping"] = o["shipping_buyer"] + o["shipping_cost"] + o["shipping_rebate"]
        o["net_settlement"] = (
            o["gross_revenue"]
            + o["total_fees"]
            + o["net_shipping"]
            + o["voucher_seller"]
            + o["voucher_platform"]
        )
        d = o["settlement_date"] or o["order_date"]
        o["month"] = _month_str(d)
        o["year"]  = d.year if d else 0
        results.append(o)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Shopee parser
# ─────────────────────────────────────────────────────────────────────────────

def parse_shopee(filepath: str, upload_batch: str = "") -> List[Dict[str, Any]]:
    """
    Sheet: 'Income'
    Row 1 (idx 0): section labels  (Order Info | Released Amount Details | ...)
    Row 2 (idx 1): sub-labels      (Order Income | Merchandise Subtotal | ...)
    Row 3 (idx 2): ACTUAL headers  (Sequence No. | Order ID | Product Name | ...)
    Data starts: row 4 (idx 3)

    Only process rows where 'View By' == 'Order' (skip 'Sku' detail rows).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Income"]

    rows = list(ws.iter_rows(values_only=True))
    # Header is at index 2 (row 3)
    headers = [str(h).strip() if h else "" for h in rows[2]]
    col = {h: i for i, h in enumerate(headers)}

    def g(row, name):
        idx = col.get(name.strip())
        return row[idx] if idx is not None and idx < len(row) else None

    # First pass: collect product info from Sku sub-rows
    sku_info: Dict[str, dict] = {}  # order_id -> {name, sku}
    for row in rows[3:]:
        if not any(row):
            continue
        view_by = str(g(row, "View By") or "").strip()
        if view_by.lower() != "sku":
            continue
        oid   = _normalize_order_id(g(row, "Order ID") or "")
        pname = str(g(row, "Product Name") or "").strip()
        pid   = str(g(row, "Product ID") or "").strip()
        if oid and oid not in sku_info and pname and pname != "-":
            sku_info[oid] = {"product_name": pname, "sku": pid}

    results = []
    for row in rows[3:]:
        if not any(row):
            continue
        view_by = str(g(row, "View By") or "").strip()
        if view_by.lower() != "order":
            continue

        order_id    = _normalize_order_id(g(row, "Order ID") or "")
        order_date  = _to_date(g(row, "Order Creation Date"))
        settle_date = _to_date(g(row, "Payout Completed Date"))
        d           = settle_date or order_date

        gross_rev   = _to_float(g(row, "Product Price"))
        net_settle  = _to_float(g(row, "Total Released Amount (RM)"))
        commission  = _to_float(g(row, "Commission Fee (incl. SST)"))
        service_fee = _to_float(g(row, "Service Fee (Incl. SST)"))
        txn_fee     = _to_float(g(row, "Transaction Fee (Incl. SST)"))
        ams_fee     = _to_float(g(row, "AMS Commission Fee"))
        saver_fee   = _to_float(g(row, "Saver Programme Fee (Incl. SST)"))

        ship_buyer   = _to_float(g(row, "Shipping Fee Paid by Buyer (excl. SST)"))
        ship_cost    = _to_float(g(row, "Shipping Fee Charged by Logistic Provider"))
        ship_rebate  = _to_float(g(row, "Shipping Rebate From Shopee"))
        ship_reverse = _to_float(g(row, "Reverse Shipping Fee"))
        net_ship     = ship_buyer + ship_cost + ship_rebate + ship_reverse

        voucher_seller   = _to_float(g(row, "Voucher Sponsored by Seller"))
        cofund_seller    = _to_float(g(row, "Cofund Voucher Sponsored by Seller"))
        coin_seller      = _to_float(g(row, "Coin Cashback Sponsored by Seller"))
        total_voucher_s  = voucher_seller + cofund_seller + coin_seller

        total_fees = commission + service_fee + txn_fee + ams_fee + saver_fee

        info = sku_info.get(order_id, {})
        results.append({
            "platform":        "shopee",
            "order_id":        order_id,
            "order_date":      order_date,
            "settlement_date": settle_date,
            "month":           _month_str(d),
            "year":            d.year if d else 0,
            "product_name":    info.get("product_name"),
            "sku":             info.get("sku"),
            "qty":             1,  # enriched below via API
            "gross_revenue":   gross_rev,
            "customer_paid":   _to_float(g(row, "Amount Paid By Buyer")),
            "commission_fee":  commission,
            "transaction_fee": txn_fee,
            "service_fee":     service_fee,
            "other_fees":      ams_fee + saver_fee,
            "total_fees":      total_fees,
            "shipping_buyer":  ship_buyer,
            "shipping_cost":   ship_cost,
            "shipping_rebate": ship_rebate,
            "net_shipping":    net_ship,
            "voucher_seller":  total_voucher_s,
            "voucher_platform":0.0,
            "net_settlement":  net_settle,
            "upload_batch":    upload_batch,
        })

    # ── Enrich quantities via Shopee API ──────────────────────────────────────
    order_ids = [r["order_id"] for r in results if r["order_id"]]
    if order_ids:
        print(f"[Financials] Fetching Shopee order quantities for {len(order_ids)} orders...")
        qtys = _shopee_fetch_order_qtys(order_ids)
        for r in results:
            if r["order_id"] in qtys:
                r["qty"] = qtys[r["order_id"]]

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Auto-detect parser
# ─────────────────────────────────────────────────────────────────────────────

def detect_and_parse(filepath: str, upload_batch: str = "") -> List[Dict[str, Any]]:
    """Auto-detect platform from filename or sheet names and parse accordingly."""
    lower = filepath.lower()
    if "tiktok" in lower or "tiktokshop" in lower:
        return parse_tiktok(filepath, upload_batch)
    elif "lazada" in lower:
        return parse_lazada(filepath, upload_batch)
    elif "shopee" in lower or "income_released" in lower:
        return parse_shopee(filepath, upload_batch)

    # Try detecting from sheet names
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    if "Order details" in sheets:
        return parse_tiktok(filepath, upload_batch)
    elif "Income Overview" in sheets:
        return parse_lazada(filepath, upload_batch)
    elif "Income" in sheets and "Summary" in sheets:
        return parse_shopee(filepath, upload_batch)

    raise ValueError(f"Cannot detect platform for file: {filepath}")